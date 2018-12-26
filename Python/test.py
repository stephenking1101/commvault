#!/usr/bin/python
# coding=utf-8

# -*- coding: utf-8 -*-
import HTMLParser
import openpyxl


def load_excel(filename):
    #打开一个workbook
    wb = openpyxl.load_workbook(filename, read_only=True)

    #获取当前活跃的worksheet,默认就是第一个worksheet
    #ws = wb.active

    #当然也可以使用下面的方法

    #获取所有表格(worksheet)的名字
    sheets = wb.sheetnames

    #第一个表格的名称
    sheet_first = sheets[0]

    #获取特定的worksheet
    ws = wb[sheet_first]

    #获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    #columns = ws.columns

    #confs = {}
    #迭代所有的行
    #for row in rows:
    #    line = [col.value for col in row]
    #    if isinstance(line[1], unicode):
    #        print(line[1].encode("utf-8"))
    #        print(line[1].encode("utf-8") == "二")
    #        print(line[1] == u"二")
    #        print(line[1] == "二".decode('utf-8'))
    #    print(line)

        # for i in range(line):
        #    confs[] =

    # 通过坐标读取值
    # print(ws.cell('A1').value)    # A表示列,1表示行print ws.cell(row=1, column=1).value
    #print(ws['A1'].value)

    #html_parser = HTMLParser.HTMLParser()
    print("---------------------------------------------------------------------")
    for i, row in enumerate(rows):
        for j, col in enumerate(row):
            print(col)
            print(col.value)
    #        #col.value = html_parser.unescape(
    #        #"2017&#24180;09&#26376;27&#26085; &#26032;&#21150;&#20844;&#33258;&#21160;&#21270;&#31649;&#29702;&#24179;&#21488; GFClusterFS ")
    #        #cell = ws.cell(row=i, column=j)
    #        #cell.value = ""

    #wb.save(filename)

load_excel("C:/Users/stephenwang/Downloads/test.xlsx")

html_parser = HTMLParser.HTMLParser()
txt = html_parser.unescape("2017&#24180;09&#26376;27&#26085; &#26032;&#21150;&#20844;&#33258;&#21160;&#21270;&#31649;&#29702;&#24179;&#21488; GFClusterFS ")
print(txt)

