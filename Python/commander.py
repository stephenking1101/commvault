#!/usr/bin/python
# coding=utf-8

# -*- coding: utf-8 -*-
import argparse
from com.commvault.api import API
from com.commvault.rest_api_json import RestAPI
from com.commvault.logger import logger
from openpyxl import load_workbook
import os
import yaml
import json


class Commander:

    def __init__(self, system_config="./system_config.yaml"):
        self.logger = logger.getlogger(Commander.__name__)
        self.logger.debug("Init application with config file: %s", system_config)
        self.system_properties = {}
        if os.path.exists(system_config):
            with open(system_config, "r") as f:
                self.system_properties = yaml.safe_load(f.read())
        # print(self.system_properties.get("setup").get("field_mapping"))

        self.conf_array = []
        self.logger.debug("Inited application with configs: %s", self.system_properties)

    def create_nas_subclient(self, api=API()):
        api.login(self.system_properties.get("server"))
        template_path = "./create_subclient_template.json"
        template = {}
        if os.path.exists(template_path):
            with open(template_path, "r") as f:
                template = json.load(f)
        self.logger.debug("Create NAS subclient with template: %s", template)

        for conf in self.conf_array:
            self.logger.debug("Creating subclient in client: %s", conf.get(self.get_field_mapping_val("clientname")))

            osname = conf.get(self.get_field_mapping_val("os")).strip()
            important_system_flag = conf.get(self.get_field_mapping_val("important_system_flag")).strip()
            important_data_flag = conf.get(self.get_field_mapping_val("important_data_flag")).strip()
            datatype = conf.get(self.get_field_mapping_val("datatype")).strip()
            appname = ""
            storage_policy_name = ""
            if osname.lower() == "nas":
                appname = "File System"

            if important_system_flag == u"否" or important_data_flag == u"否":
                storage_policy_name = "L34_"
            else:
                storage_policy_name = "L12_"

            if datatype == u"应用日志":
                storage_policy_name = storage_policy_name + "APPLOG_31D_I_10Y_1"

            api.createsubclient(appname=appname,
                                clientname='win-2012-server',
                                subclientname=storage_policy_name,
                                storage_policy_name="SP-FS-1D",
                                paths=["C:\\temp"],
                                descpt=conf.get(self.get_field_mapping_val("description")).strip(),
                                content_operation_type="ADD",
                                template=template)

        api.logout()

    def get_field_mapping_val(self, key):
        mapping = self.system_properties.get("setup").get("field_mapping")
        return mapping.get(key)

    def load_excel(self, filename):
        self.logger.debug("Going to load excel file: %s", filename)
        # 打开一个workbook
        wb = load_workbook(filename, read_only=True)

        # 获取所有表格(worksheet)的名字
        sheets = wb.sheetnames

        # 第一个表格的名称
        sheet_first = sheets[0]

        # 获取特定的worksheet
        ws = wb[sheet_first]

        # 获取表格所有行和列，两者都是可迭代的
        rows = ws.rows

        col_title = []
        # 迭代所有的行
        for i, row in enumerate(rows):
            conf = {}
            for j, col in enumerate(row):
                if i == 0:
                    if j == 0 and col.value is None:
                        raise RuntimeError("无法读取数据，首行必须为标题行")
                    else:
                        col_title.append(col.value)
                else:
                    conf[col_title[j]] = col.value

            if len(conf) > 0:
                self.conf_array.append(conf)

        self.logger.debug("Loaded excel data: %s", self.conf_array)
        return self.conf_array


# sub-command functions
def create_subclient(args, api=API()):
    api.login(args.server)

    if args.path is None:
        paths = []
    else:
        paths = args.path

    api.createsubclient(args.appName, args.clientName, args.subclientName, args.storagePolicyName, paths,
                        args.numberOfBackupStreams, args.backupsetName, args.description, args.contentOperationType)
    api.logout()


def cmd():
    applogger = logger.getlogger(__name__)

    # create top-level parser
    parser = argparse.ArgumentParser()

    # 定位参数的使用, python cli_fw.py 8
    parser.add_argument("server", help="web server ip address/host name", type=str)
    parser.add_argument("userName", help="user to login the CommServe", type=str)
    parser.add_argument("password", help="password of the specific user", type=str)

    subparsers = parser.add_subparsers(help="sub-command help")

    # create the parser for the 'createSubclient' command
    parser_create_subclient = subparsers.add_parser('createSubclient', help='create subclient')
    parser_create_subclient.set_defaults(func=create_subclient)
    parser_create_subclient.add_argument('appName', type=str, help='agent type')
    parser_create_subclient.add_argument('clientName', type=str, help='client name')
    parser_create_subclient.add_argument('subclientName', type=str, help='subclient name')
    parser_create_subclient.add_argument('--numberOfBackupStreams', type=int, help='number of backup streams',
                                         default=1)
    parser_create_subclient.add_argument('--description', type=str, help='subclient description')
    parser_create_subclient.add_argument('--storagePolicyName', type=str, help='storage policy to be associated')
    parser_create_subclient.add_argument('--backupsetName', type=str, help='backupset name', default="defaultBackupSet")
    parser_create_subclient.add_argument('--contentOperationType', type=str, help='add/del/modify content',
                                         default="ADD")
    parser_create_subclient.add_argument('--path', action='append', type=str,
                                         help='content to backup, can be more then one --path')

    args = parser.parse_args()
    applogger.debug("arguments: %s", args)
    # for test
    # parser.parse_args(['--help'])
    # parser.parse_args(['createSubclient', '-h'])
    # args = parser.parse_args("server user pwd createSubclient app client subclient".split())

    api = RestAPI(args.userName, args.password)
    # call sub command
    args.func(args, api)


if __name__ == "__main__":
    # main()
    parser = argparse.ArgumentParser()
    parser.add_argument("userName", help="user to login the CommServe", type=str)
    parser.add_argument("password", help="password of the specific user", type=str)

    args = parser.parse_args()
    api = RestAPI(args.userName, args.password)

    c = Commander()
    c.load_excel("C:/Users/stephenwang/Downloads/clientconfig.xlsx")
    c.create_nas_subclient(api)
