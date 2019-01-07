#!/usr/bin/python
# coding=utf-8

# -*- coding: utf-8 -*-
import argparse
from commvault.api import API
from commvault.rest_api_json import RestAPI
from commvault.logger import logger
from commvault.util import dict_util
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import os
import yaml
import json


class Commander:

    def __init__(self, system_config="./system_config.yaml"):
        self.logger = logger.getlogger(Commander.__name__)
        self.logger.debug("Init application with config file: %s", system_config)
        self.system_properties = {}
        if os.path.exists(system_config):
            with open(system_config, "r") as f:
                self.system_properties = yaml.safe_load(f.read())
        # print(self.system_properties.get("setup").get("field_mapping"))

        self.conf_array = []
        self.logger.debug("Inited application with configs: %s", self.system_properties)

    def create_subclients(self, api=API()):
        src_path = os.path.dirname(os.path.realpath(__file__))
        # cwd = os.getcwd() # current working directory

        api.login(self.system_properties.get("server"))
        template_path = src_path + "/commvault/template/create_subclient_template.json"
        template = {}
        if os.path.exists(template_path):
            with open(template_path, "r") as f:
                template = json.load(f)
        self.logger.debug("Create subclients with template: %s", template)

        schedule_policy_template = {}
        if os.path.exists(src_path + "/commvault/template/update_schedule_policy_entity_assoc_template.json"):
            with open(src_path + "/commvault/template/update_schedule_policy_entity_assoc_template.json", "r") as f:
                schedule_policy_template = json.load(f)
        self.logger.debug("Create subclients with schedule policy template: %s", schedule_policy_template)

        storage_policy_name_list, storage_policies = api.get_storage_policy()
        storage_policy_name_list.sort(reverse=True)

        schedule_policy_name_list, schedule_policies = api.get_schedule_policy()

        str_func = lambda val: val if val and isinstance(val, basestring) else (str(val) if val else "")
        for conf in self.conf_array:
            self.logger.debug("Creating subclient in client: %s", conf.get(self.get_field_mapping_val("clientname")))

            osname = dict_util.dict_get_converted_value(conf, self.get_field_mapping_val("os"), str_func).strip()
            important_system_flag = dict_util.dict_get_converted_value(conf, self.get_field_mapping_val(
                "important_system_flag"), str_func).strip()
            important_data_flag = dict_util.dict_get_converted_value(conf, self.get_field_mapping_val(
                "important_data_flag"), str_func).strip()
            datatype = dict_util.dict_get_converted_value(conf, self.get_field_mapping_val("datatype"), str_func).strip()
            clientname = dict_util.dict_get_converted_value(conf, self.get_field_mapping_val("clientname"), str_func).strip()
            database_instance_name = dict_util.dict_get_converted_value(conf, self.get_field_mapping_val(
                "database_instance_name"), str_func).strip()
            database_name = dict_util.dict_get_converted_value(conf,
                                                               self.get_field_mapping_val("database_name"), str_func).strip()
            database_installation_path = dict_util.dict_get_converted_value(conf, self.get_field_mapping_val(
                "database_installation_path"), str_func).strip()
            start_time = dict_util.dict_get_converted_value(conf, self.get_field_mapping_val(
                "start_time"), str_func).strip()
            start_time = start_time[:start_time.index(":")+3]

            storage_policy_name = self.get_storage_policy_name(important_system_flag, important_data_flag, datatype,
                                                               storage_policy_name_list)

            subclient_list, subclient_properties = api.get_subclient(client_name=clientname)
            subclient_name = self.get_subclient_name(storage_policy_name, subclient_list)

            schedule_policy_name = self.get_schedule_policy_name(osname.lower(), datatype, start_time, schedule_policy_name_list)

            subclient_id = self.do_create_subclient(osname=osname.lower(),
                                 datatype=datatype,
                                 clientname=clientname,
                                 subclient_name=subclient_name,
                                 storage_policy_name=storage_policy_name,
                                 paths=[dict_util.dict_get_converted_value(conf,
                                                                           self.get_field_mapping_val("path"), str_func).strip()],
                                 database_instance_name=database_instance_name,
                                 database_name=database_name,
                                 database_installation_path=database_installation_path,
                                 descpt=dict_util.dict_get_converted_value(conf, self.get_field_mapping_val(
                                     "description"), str_func).strip(),
                                 template=template,
                                 api=api)

            if subclient_id:
                api.update_schedule_policy_entry_assoc(template=schedule_policy_template, subclientId=subclient_id, taskName=schedule_policy_name)

        api.logout()

    def create_oracle_instance(self, command, api=API()):
        src_path = os.path.dirname(os.path.realpath(__file__))
        template = ET.parse(src_path + "/commvault/template/CreateInstance_Template.xml")
        r = api.post_execute_qcommand(ET.tostring(template.getroot()), command)
        self.logger.debug("Response from server: %s", r.json() if r else r)

    def update_oracle_subclient(self, command, api=API()):
        src_path = os.path.dirname(os.path.realpath(__file__))
        template = ET.parse(src_path + "/commvault/template/update_subclient_template.xml")
        r = api.post_execute_qcommand(ET.tostring(template.getroot()), command)
        self.logger.debug("Response from server: %s", r.json() if r else r)

    def get_oracle_subclient_id_from_list(self, subclientName, instanceName, list):
        subclient_id = None;
        for subclient in list:
            if subclient.get("subClientEntity") and subclient.get("subClientEntity").get("appName").lower() == "oracle":
                if subclient.get("subClientEntity").get("subclientName") == subclientName and subclient.get("subClientEntity").get("instanceName") == instanceName:
                    subclient_id = subclient.get("subClientEntity").get("subclientId")
                    break

        return subclient_id

    def do_create_subclient(self, osname, datatype, clientname,
                            subclient_name, storage_policy_name, paths,
                            database_instance_name, database_name, database_installation_path,
                            descpt, template, api=API()):
        if osname == "nas":
            appname = "NAS"
            return api.create_subclient(appname=appname,
                                 clientname=clientname,
                                 subclientname=subclient_name,
                                 storage_policy_name=storage_policy_name,
                                 paths=paths,
                                 descpt=descpt,
                                 content_operation_type="ADD",
                                 template=template)

        elif osname == "linux" or osname == "windows":
            if datatype == "Oracle":
                command = "qoperation execute -clientName " + clientname + " -instanceName " + database_instance_name + " -logBackupStoragePolicy/storagePolicyName " + storage_policy_name + " -commandLineStoragePolicy/storagePolicyName " + storage_policy_name + " -dataArchiveGroup/storagePolicyName " + storage_policy_name + " -oracleHome '" + database_installation_path + "' -oracleUser/userName oracle -sqlConnect/userName '/' -useCatalogConnect false"
                self.create_oracle_instance(command, api)

                # rename oracle subclient
                command = "qoperation execute -clientName " + clientname + " -instanceName " + database_instance_name + " -subclientName 'default' -newName " + storage_policy_name
                self.update_oracle_subclient(command, api)

                # update oracle subclient properties
                command = "qoperation execute -appName 'Oracle' -clientName " + clientname + " -instanceName " + database_instance_name + " -subclientName " + storage_policy_name + " -selectiveOnlineFull true -dataBackupStoragePolicy/storagePolicyName " + storage_policy_name + " -softwareCompression 'USE_STORAGE_POLICY_SETTINGS' -backupSPFile true -dataFilesPerBFS 1 -skipInaccessible true -archiveDelete true"
                self.update_oracle_subclient(command, api)

                # update oracle subclient properties
                command = "qoperation execute -appName 'Oracle' -clientName " + clientname + " -instanceName " + database_instance_name + " -subclientName 'ArchiveLog' -softwareCompression 'USE_STORAGE_POLICY_SETTINGS' -backupControlFile false -archiveDelete false"
                self.update_oracle_subclient(command, api)

                subclient_list, subclient_properties = api.get_subclient(client_name=clientname)

                return self.get_oracle_subclient_id_from_list(storage_policy_name, database_instance_name, subclient_properties)
        else:
            return None

    def get_field_mapping_val(self, key):
        mapping = dict_util.dict_get_child(self.system_properties, ("setup", "field_mapping"))
        return mapping.get(key)

    def get_storage_policy_name(self, important_system_flag, important_data_flag, datatype, storage_policy_name_list):
        if important_system_flag == u"否" or important_data_flag == u"否":
            storage_policy_name = "L34_"
        else:
            storage_policy_name = "L12_"

        if datatype == u"应用日志":
            storage_policy_name = storage_policy_name + "APPLOG_"
        elif datatype == "Oracle":
            storage_policy_name = storage_policy_name + "DB_"

        for sp in storage_policy_name_list:
            if sp.startswith(storage_policy_name):
                storage_policy_name = sp
                break

        self.logger.debug("Got storage policy: %s", storage_policy_name)
        return storage_policy_name

    def get_schedule_policy_name(self, osname, datatype, start_time, schedule_policy_name_list):
        self.logger.debug("Schedule start time: %s", start_time)
        if osname == "nas":
            schedule_policy_name = "NDMP_MonthlyFull"
        elif datatype == u"应用日志":
            schedule_policy_name = "FS_MonthlyFull"
        elif datatype == "Oracle":
            schedule_policy_name = "DB_DailyFull"

        schedule_list = [sp for sp in schedule_policy_name_list if sp.startswith(schedule_policy_name)]

        for sp in schedule_list:
            if start_time in sp:
                schedule_policy_name = sp
                break

        self.logger.debug("Got schedule policy: %s", schedule_policy_name)
        return schedule_policy_name

    def get_subclient_name(self, propose_name, name_list):
        subclient_name = propose_name
        while subclient_name in name_list:
            self.logger.debug("Subclient %s already exists", subclient_name)
            seq = subclient_name[subclient_name.rindex("_") + 1:]
            subclient_name = subclient_name[:subclient_name.rindex("_") + 1] + str(int(seq) + 1)

        return subclient_name

    def load_excel(self, filename):
        self.logger.debug("Going to load excel file: %s", filename)
        # 打开一个workbook
        wb = load_workbook(filename, read_only=True)

        # 获取所有表格(worksheet)的名字
        sheets = wb.sheetnames

        # 第一个表格的名称
        sheet_first = sheets[0]

        # 获取特定的worksheet
        ws = wb[sheet_first]

        # 获取表格所有行和列，两者都是可迭代的
        rows = ws.rows

        col_title = []
        # 迭代所有的行
        for i, row in enumerate(rows):
            conf = {}
            for j, col in enumerate(row):
                if i == 0:
                    if j == 0 and col.value is None:
                        raise RuntimeError("无法读取数据，首行必须为标题行")
                    else:
                        col_title.append(col.value)
                else:
                    conf[col_title[j]] = col.value

            if len(conf) > 0:
                self.conf_array.append(conf)

        self.logger.debug("Loaded excel data: %s", self.conf_array)
        return self.conf_array


def main():
    applogger = logger.getlogger(__name__)
    src_path = os.path.dirname(os.path.realpath(__file__))
    com = Commander(system_config=src_path + "/system_config.yaml")

    # create top-level parser
    parser = argparse.ArgumentParser()

    # 定位参数的使用, python cli_fw.py 8
    parser.add_argument("userName", help="user to login the CommServe", type=str)
    parser.add_argument("password", help="password of the specific user", type=str)

    subparsers = parser.add_subparsers(help="batchJob -h to show help")

    # create the parser for the 'batchJob' command
    parser_create_subclient = subparsers.add_parser('batchJob', help='Bulk client/subclient creation and configuration')
    parser_create_subclient.set_defaults(func=com.create_subclients)
    parser_create_subclient.add_argument('filepath', type=str,
                                         help='Excel file path that contains subclient config')

    args = parser.parse_args()
    applogger.debug("arguments: %s", args)
    # for test
    # parser.parse_args(['--help'])
    # parser.parse_args(['createSubclient', '-h'])
    # args = parser.parse_args("server user pwd createSubclient app client subclient".split())

    api = RestAPI(args.userName, args.password)
    com.load_excel(args.filepath)
    # call sub command
    args.func(api)


if __name__ == "__main__":
    main()
